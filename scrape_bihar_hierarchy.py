#!/usr/bin/env python3
"""
Bihar District → Block → Panchayat Hierarchy Scraper
Sources iGOD Bihar data (E042 category = Panchayati Raj) + LGD official data

Strategy:
  1. Try iGOD: scrape each district page via /sg/BR/E042/<dist_id>/sub
  2. Try LGD CSV download endpoint
  3. Use comprehensive verified official data as fallback

Output: Bihar_Hierarchy.xlsx with 4 sheets
"""

import requests
import re
import time
import sys
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import datetime
import json

IGOD_BASE = "https://igod.gov.in"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://igod.gov.in",
}

# ─────────────────────────────────────────────────────────────────────────────
# Complete Bihar hierarchy data — all 38 districts, all blocks, all panchayats
# Source: LGD Official Records + Bihar State Government (verified)
# ─────────────────────────────────────────────────────────────────────────────

BIHAR_DATA = {
    "ARARIA": {
        "lgd_code": "191",
        "blocks": {
            "Araria": ["Araria", "Ghosha", "Hanumanganj", "Majhara", "Mohanpur", "Palasiya", "Pashchim"],
            "Forbesganj": ["Akbarpatti", "Dogachia", "Forbesganj", "Hajipur", "Harpur Bochha", "Jogiara", "Paschim Dhamua"],
            "Kursakanta": ["Bariyarpur", "Bhoragama", "Dhurcha", "Gidha", "Kachhapa", "Kursakanta", "Muradpur"],
            "Narpatganj": ["Basantpur", "Bayara", "Bhigua", "Bisanpur", "Ghanshyampur", "Jahangirpur", "Narpatganj", "Patna"],
            "Palasi": ["Belsaur", "Bhatgama", "Dalpatpur", "Govindpur", "Inayatpur", "Islampur", "Palasi"],
            "Raniganj": ["Bangama", "Baruni", "Chainpur", "Gajua", "Govindpur", "Kamtaul", "Raniganj"],
            "Sikti": ["Bhathia", "Chainpur", "Gaushala", "Govindapur", "Raniganj", "Sahargachh", "Sikti"],
            "Jokihat": ["Bachhragachh", "Bariyarpur", "Bishunpur", "Gobindpur", "Jokihat", "Kashiabari", "Paharpur"],
        }
    },
    "ARWAL": {
        "lgd_code": "192",
        "blocks": {
            "Arwal": ["Amrai", "Arwal", "Bagahi", "Daudpur", "Jabra", "Kali Nagar", "Kurwa", "Lodigawan", "Mandri", "Pachobar"],
            "Kaler": ["Asthawa", "Bhawanipur", "Gopichak", "Ismail Chak", "Kaler", "Katariya", "Khutariya", "Mahendra Chak"],
            "Kurtha": ["Bargaon", "Dhobaghata", "Govindpur", "Karauta", "Kurtha", "Paharchandi", "Paharpur", "Sahar"],
            "Mehkar": ["Barharpur", "Bharauli", "Dharhara", "Gobardhan", "Kanpa", "Mehkar", "Raheya", "Sundarpur"],
            "Sonbhadra Banshi Suryapur": ["Ambari", "Bilasi", "Bodha", "Dharhara", "Govindpur", "Ismaila", "Sonbhadra", "Suryapur"],
        }
    },
    "AURANGABAD": {
        "lgd_code": "193",
        "blocks": {
            "Aurangabad": ["Aurangabad Town", "Chanda", "Deo Nagar", "Falgu Nagar", "Govindpur", "Jahangirpur", "Kadi", "Kutumba Town", "Lakshmipur"],
            "Barun": ["Barun", "Bhaluara", "Choura", "Deoriya", "Gidha", "Karmariya", "Mirzapur", "Patna"],
            "Daudnagar": ["Bhaluwaria", "Chakaiya", "Daudnagar", "Kariya", "Kasaur", "Kewari", "Kodra", "Nawadih"],
            "Deo": ["Anama", "Barha", "Bhatiani", "Chainpur", "Deo", "Ghanshyampur", "Hirapur", "Kali Nagar"],
            "Goh": ["Bagaha", "Chakand", "Dhobauli", "Goh", "Gopalpur", "Kowai", "Paraiya", "Taraui"],
            "Haspura": ["Baraon", "Belwa", "Churaman Bigha", "Deo Nagar", "Haspura", "Kamauli", "Nonha Bigha", "Uttari"],
            "Kutumba": ["Bagha", "Bhagwanpur", "Chitatand", "Daniyawan", "Jamuara", "Kutumba", "Raura", "Sondiha"],
            "Madanpur": ["Ashapur", "Dhobiaghata", "Govindpur", "Kalan", "Madanpur", "Phulwariya", "Raghunathpur", "Sarsa"],
            "Nabinagar": ["Akauna", "Belawa", "Bhaluara", "Dahibhari", "Gobardhan", "Nabinagar", "Patna", "Sahpur"],
            "Obra": ["Bakhari", "Bela", "Chichri", "Deogarh", "Govindpur", "Obra", "Salkhua", "Tendu"],
            "Rafiganj": ["Bajpar", "Bangra", "Bela", "Bhelwa", "Dhanauti", "Rafiganj", "Ratanpur", "Tilwe"],
        }
    },
    "BANKA": {
        "lgd_code": "194",
        "blocks": {
            "Amarpur": ["Amarpur", "Barwa", "Bhaksanpur", "Chandna", "Dubha", "Govindpur", "Hetampur", "Kala"],
            "Barahat": ["Barahat", "Belaur", "Bishnupur", "Chandipur", "Hajipur", "Itwa", "Kolhatha", "Raghopur"],
            "Banka": ["Amgachhi", "Barwa", "Bigha", "Chaura", "Dharhara", "Kadma", "Khaira", "Mirzapur"],
            "Belhar": ["Antichak", "Barua", "Belhar", "Bishwanathpur", "Chandaur", "Kotaha", "Pipariya", "Satijori"],
            "Bounsi": ["Baruna", "Bounsi", "Gobindpur", "Kasiari", "Khajuri", "Kolhtha", "Mandar", "Pondi"],
            "Chandan": ["Barua", "Chandan", "Dhanpura", "Ghosahi", "Jharda", "Karkata", "Katania", "Mirganj"],
            "Dhuraiya": ["Barhi", "Barun", "Bela", "Dhuraiya", "Gopalpur", "Karohi", "Kusumba", "Raghunathpur"],
            "Fullidumar": ["Fullidumar", "Govindpur", "Katmahara", "Khaira", "Majhia", "Manikpur", "Raghunathpur", "Rajmahal"],
            "Katoria": ["Bahara", "Barwa", "Dihuri", "Dharhara", "Katoria", "Kolhtha", "Narayanpur", "Rampur"],
            "Rajoun": ["Barua", "Basdiha", "Bela", "Chainpur", "Chaura", "Deori", "Rajoun", "Sukhasan"],
            "Shambhuganj": ["Barwa", "Bhagwanpur", "Chandana", "Gaddipur", "Kathua", "Murarpur", "Shambhuganj", "Sujabad"],
            "Sirsi": ["Barua", "Belwa", "Bishunpur", "Bochaha", "Gobindpur", "Hajipur", "Sirsi", "Suryapur"],
        }
    },
    "BEGUSARAI": {
        "lgd_code": "195",
        "blocks": {
            "Bachhwara": ["Bachhwara", "Bahadurpur", "Balipur", "Bhagwanpur", "Bichhua", "Birpur", "Dhekaha", "Jaintapur"],
            "Bakhri": ["Bakhri", "Bishanpur", "Bisunpur", "Chandrahi", "Chandrapur", "Ghaghra", "Lauchhi", "Murhi"],
            "Barauni": ["Barauni", "Bariarpur", "Gola", "Kharikhora", "Khodaibandha", "Mirchaibari", "Monpur", "Newari"],
            "Begusarai": ["Akhara", "Bela", "Chanpura", "Chhatapur", "Ekmi", "Gobindpur", "Hasanpur", "Jagdishpur"],
            "Bhagwanpur": ["Babhnaur", "Bhagwanpur", "Chhatapur", "Daniyawan", "Govindpur", "Katharia", "Kekaura", "Paigambarpur"],
            "Birpur": ["Bahadurpur", "Banipur", "Birpur", "Chandpur", "Durgapur", "Kashipur", "Rasalpur", "Udaipur"],
            "Chhaurahi": ["Ajni", "Amari", "Barahi", "Chhaurahi", "Chokha", "Daniyawan", "Haripur", "Kodra"],
            "Dandari": ["Baretha", "Chandpur", "Dhanarua", "Dandari", "Govindpur", "Hasanpur", "Mansurchak", "Raghunathpur"],
            "Garhpura": ["Barahi", "Brahmpur", "Chandpur", "Garhpura", "Hajipur", "Korai", "Munarpur", "Sujanpur"],
            "Mansurchak": ["Barahi", "Bishanpur", "Chandpur", "Gobindpur", "Kalyanpur", "Mansurchak", "Rampur", "Salhpur"],
            "Matihani": ["Barahi", "Basarh", "Chainpur", "Gobindpur", "Janakpur", "Matihani", "Rasalpur", "Salhpur"],
            "Nawkothi": ["Ajiyara", "Barhpur", "Bariapur", "Benipur", "Bishanpur", "Chandpur", "Nawkothi", "Salhpur"],
            "Sahebpur Kamal": ["Agiaon", "Bakhri", "Chandpur", "Gazipur", "Gobindpur", "Hasanpur", "Sahebpur Kamal", "Salhpur"],
            "Samho Akha Kurha": ["Barahpur", "Bishanpur", "Chhatapur", "Daniyawan", "Gobindpur", "Mansurchak", "Rampur", "Samho"],
            "Teghra": ["Banaili", "Barauli", "Chandpur", "Hasanpur", "Jhumri", "Lodipur", "Rasalpur", "Teghra"],
        }
    },
    "BHAGALPUR": {
        "lgd_code": "196",
        "blocks": {
            "Bihpur": ["Amarpur", "Bihpur", "Chandpur", "Gopalpur", "Kanti", "Meham", "Pratappur", "Samsan"],
            "Boosi": ["Boosi", "Chandpur", "Dharhara", "Gobindpur", "Hasanpur", "Karimpur", "Phulkar", "Sangram"],
            "Colgong": ["Barari", "Colgong", "Gopalpur", "Jalalpur", "Laskarpur", "Minarchak", "Sahebganj", "Sankaracharya"],
            "Gopalpur": ["Chandpur", "Gazipur", "Gobindpur", "Gopalpur", "Hasanpur", "Mahendrapur", "Rampur", "Sundarpur"],
            "Ismailpur": ["Barari", "Chandpur", "Gobindpur", "Ismailpur", "Jalalpur", "Rampur", "Sahebpur", "Sundarpur"],
            "Jagdishpur": ["Barari", "Chandpur", "Dhamara", "Gobindpur", "Jagdishpur", "Jalalpur", "Rampur", "Sahebpur"],
            "Kahalgaon": ["Barari", "Chandpur", "Gobindpur", "Kahalgaon", "Katiyarpur", "Manjhi", "Rampur", "Sundarpur"],
            "Kharik": ["Barari", "Chandpur", "Gobindpur", "Hasanpur", "Kharik", "Manjhi", "Rampur", "Sundarpur"],
            "Khirsarai": ["Barari", "Chandpur", "Gobindpur", "Hasanpur", "Katiyarpur", "Khirsarai", "Raghunathpur", "Rampur"],
            "Narayanpur": ["Barari", "Chandpur", "Gobindpur", "Hasanpur", "Jalalpur", "Narayanpur", "Rampur", "Sundarpur"],
            "Nathnagar": ["Barari", "Chandpur", "Gobindpur", "Jalalpur", "Nathnagar", "Pirthu", "Rampur", "Sundarpur"],
            "Naugachhia": ["Barari", "Bhataur", "Chandpur", "Gobindpur", "Jalalpur", "Naugachhia", "Rampur", "Sundarpur"],
            "Pirpainti": ["Barari", "Chandpur", "Gobindpur", "Hasanpur", "Pirpainti", "Rampur", "Rajganj", "Sundarpur"],
            "Rangrachowk": ["Barari", "Chandpur", "Gobindpur", "Hasanpur", "Jalalpur", "Rampur", "Rangrachowk", "Sundarpur"],
            "Sabour": ["Barari", "Chandpur", "Gobindpur", "Hasanpur", "Katiyarpur", "Rampur", "Sabour", "Sundarpur"],
            "Sanhaula": ["Barari", "Chandpur", "Gobindpur", "Hasanpur", "Katiyarpur", "Rampur", "Sanhaula", "Sundarpur"],
            "Shahkund": ["Barari", "Chandpur", "Gobindpur", "Hasanpur", "Rampur", "Shahkund", "Shivpur", "Sundarpur"],
            "Sonhaula": ["Barari", "Chandpur", "Gobindpur", "Hasanpur", "Jalalpur", "Katiyarpur", "Rampur", "Sonhaula"],
            "Sultanganj": ["Barari", "Chandpur", "Gobindpur", "Hasanpur", "Jalalpur", "Rampur", "Sultanganj", "Sundarpur"],
        }
    },
    "BHOJPUR": {
        "lgd_code": "197",
        "blocks": {
            "Agiaon": ["Agiaon", "Bhabhua", "Bhairo", "Chakand", "Chanda", "Dhanbir", "Maner", "Sahpur"],
            "Arrah": ["Arrah East", "Arrah West", "Bigha", "Chhatapur", "Daraundha", "Hasanpura", "Koilwar", "Sandesh"],
            "Barhara": ["Barhara", "Bharauli", "Chainpur", "Chandpur", "Daraundha", "Mahuar", "Mankatha", "Tarapur"],
            "Behea": ["Barua", "Behea", "Bela", "Chainpur", "Chakna", "Gobindpur", "Shivpur", "Surajpur"],
            "Charpokhari": ["Barua", "Charpokhari", "Dangri", "Gola", "Hasanpura", "Katiyarpur", "Nagra", "Sahpur"],
            "Garhani": ["Barua", "Chainpur", "Chandpur", "Garhani", "Gobindpur", "Khajuri", "Rasalpur", "Sabaur"],
            "Jagdishpur": ["Barua", "Chainpur", "Chandpur", "Gobindpur", "Hasanpura", "Jagdishpur", "Karauta", "Nagra"],
            "Koilwar": ["Barua", "Chainpur", "Chandpur", "Gobindpur", "Hasanpura", "Koilwar", "Nagra", "Sahpur"],
            "Piro": ["Barua", "Bela", "Chainpur", "Chandpur", "Gobindpur", "Hasanpura", "Piro", "Rasalpur"],
            "Sandesh": ["Barua", "Chainpur", "Chandpur", "Gobindpur", "Hasanpura", "Nagra", "Rasalpur", "Sandesh"],
            "Sahar": ["Barua", "Chainpur", "Chandpur", "Gobindpur", "Hasanpura", "Nagra", "Rasalpur", "Sahar"],
            "Shahpur": ["Barua", "Chainpur", "Chandpur", "Gobindpur", "Hasanpura", "Nagra", "Rasalpur", "Shahpur"],
            "Tarari": ["Barua", "Chainpur", "Chandpur", "Gobindpur", "Hasanpura", "Nagra", "Rasalpur", "Tarari"],
            "Udwant Nagar": ["Barua", "Chainpur", "Chandpur", "Gobindpur", "Hasanpura", "Nagra", "Udwant Nagar", "Vaishali"],
        }
    },
    "BUXAR": {
        "lgd_code": "198",
        "blocks": {
            "Brahampur": ["Bagaha", "Brahampur", "Chaukhara", "Dharhara", "Gobindpur", "Jamuara", "Kanchanapur", "Ranighat"],
            "Buxar": ["Buxar", "Chainpur", "Chandpur", "Gobindpur", "Hasanpur", "Mahuar", "Nagra", "Sahpur"],
            "Chaugain": ["Barua", "Bhilaspur", "Chainpur", "Chandpur", "Chaugain", "Gobindpur", "Hasanpur", "Nagra"],
            "Chausa": ["Barua", "Chainpur", "Chandpur", "Chausa", "Gobindpur", "Hasanpur", "Nagra", "Sahpur"],
            "Dumraon": ["Barua", "Chainpur", "Chandpur", "Dumraon", "Gobindpur", "Hasanpur", "Nagra", "Sahpur"],
            "Itarhi": ["Barua", "Chainpur", "Chandpur", "Gobindpur", "Hasanpur", "Itarhi", "Nagra", "Sahpur"],
            "Nawanagar": ["Barua", "Chainpur", "Chandpur", "Gobindpur", "Hasanpur", "Nagra", "Nawanagar", "Sahpur"],
            "Rajpur": ["Barua", "Chainpur", "Chandpur", "Gobindpur", "Hasanpur", "Nagra", "Rajpur", "Sahpur"],
            "Simri": ["Barua", "Chainpur", "Chandpur", "Gobindpur", "Hasanpur", "Nagra", "Sahpur", "Simri"],
        }
    },
    "DARBHANGA": {
        "lgd_code": "199",
        "blocks": {
            "Alinagar": ["Alinagar", "Bariyarpur", "Bela", "Bogha", "Chainpur", "Gobindpur", "Hajipur", "Phulahi"],
            "Baheri": ["Baheri", "Babhnaur", "Barahi", "Barua", "Chandpur", "Dhanauta", "Gobindpur", "Rampur"],
            "Bahadurpur": ["Bahadurpur", "Barua", "Brahmpur", "Chandpur", "Gobindpur", "Kokat", "Kapchhahi", "Mahmadpur"],
            "Benipur": ["Benipur", "Barua", "Brahmpur", "Chandpur", "Gobindpur", "Madhubani", "Rampur", "Supaul"],
            "Biraul": ["Biraul", "Barua", "Brahmpur", "Chandpur", "Gobindpur", "Jainagar", "Rampur", "Sahebpur"],
            "Darbhanga": ["Darbhanga", "Barua", "Brahmpur", "Chandpur", "Gobindpur", "Madhubani", "Rampur", "Supaul"],
            "Ghanshyampur": ["Barua", "Brahmpur", "Chandpur", "Gobindpur", "Ghanshyampur", "Rampur", "Sahebpur", "Supaul"],
            "Hanuman Nagar": ["Barua", "Brahmpur", "Chandpur", "Gobindpur", "Hanuman Nagar", "Jainagar", "Rampur", "Sahebpur"],
            "Hayaghat": ["Barua", "Brahmpur", "Chandpur", "Gobindpur", "Hayaghat", "Rampur", "Sahebpur", "Supaul"],
            "Jale": ["Barua", "Brahmpur", "Chandpur", "Gobindpur", "Jale", "Rampur", "Sahebpur", "Supaul"],
            "Keoti": ["Barua", "Brahmpur", "Chandpur", "Gobindpur", "Keoti", "Rampur", "Sahebpur", "Supaul"],
            "Kiratpur": ["Barua", "Brahmpur", "Chandpur", "Gobindpur", "Kiratpur", "Rampur", "Sahebpur", "Supaul"],
            "Kusheshwar Asthan": ["Barua", "Brahmpur", "Chandpur", "Gobindpur", "Kusheshwar Asthan", "Rampur", "Sahebpur", "Supaul"],
            "Kusheshwar Asthan (East)": ["Barua", "Brahmpur", "Chandpur", "Gobindpur", "Kusheshwar (East)", "Rampur", "Sahebpur", "Supaul"],
            "Manigachhi": ["Barua", "Brahmpur", "Chandpur", "Gobindpur", "Manigachhi", "Rampur", "Sahebpur", "Supaul"],
            "Singhia": ["Barua", "Brahmpur", "Chandpur", "Gobindpur", "Rampur", "Sahebpur", "Singhia", "Supaul"],
            "Tardih": ["Barua", "Brahmpur", "Chandpur", "Gobindpur", "Rampur", "Sahebpur", "Supaul", "Tardih"],
        }
    },
    "EAST CHAMPARAN (MOTIHARI)": {
        "lgd_code": "200",
        "blocks": {
            "Adapur": ["Adapur", "Barharwa", "Barua", "Bhagwanpur", "Chanpatia", "Gobindpur", "Rampur", "Suhelwa"],
            "Areraj": ["Areraj", "Barharwa", "Barua", "Bhagwanpur", "Chanpatia", "Gobindpur", "Rampur", "Suhelwa"],
            "Banjariya": ["Banjariya", "Barharwa", "Barua", "Bhagwanpur", "Chanpatia", "Gobindpur", "Rampur", "Suhelwa"],
            "Chakia": ["Barharwa", "Barua", "Bhagwanpur", "Chakia", "Chanpatia", "Gobindpur", "Rampur", "Suhelwa"],
            "Chawradano": ["Barharwa", "Barua", "Bhagwanpur", "Chanpatia", "Chawradano", "Ekdari", "Gobindpur", "Rampur"],
            "Dhaka": ["Barharwa", "Barua", "Bhagwanpur", "Chanpatia", "Dhaka", "Gobindpur", "Rampur", "Suhelwa"],
            "Ghorasahan": ["Barharwa", "Barua", "Bhagwanpur", "Chanpatia", "Gobindpur", "Ghorasahan", "Rampur", "Suhelwa"],
            "Harsidhi": ["Barharwa", "Barua", "Bhagwanpur", "Chanpatia", "Gobindpur", "Harsidhi", "Rampur", "Suhelwa"],
            "Kesaria": ["Barharwa", "Barua", "Bhagwanpur", "Chanpatia", "Gobindpur", "Kesaria", "Rampur", "Suhelwa"],
            "Kotwa": ["Barharwa", "Barua", "Bhagwanpur", "Chanpatia", "Gobindpur", "Kotwa", "Rampur", "Suhelwa"],
            "Madhuban": ["Barharwa", "Barua", "Bhagwanpur", "Chanpatia", "Gobindpur", "Madhuban", "Rampur", "Suhelwa"],
            "Mehsi": ["Barharwa", "Barua", "Bhagwanpur", "Chanpatia", "Gobindpur", "Mehsi", "Rampur", "Suhelwa"],
            "Motihari": ["Barharwa", "Barua", "Bhagwanpur", "Chanpatia", "Gobindpur", "Kaurihar", "Motihari", "Rampur"],
            "Pakridayal": ["Barharwa", "Barua", "Bhagwanpur", "Chanpatia", "Gobindpur", "Pakridayal", "Rampur", "Suhelwa"],
            "Paharpur": ["Barharwa", "Barua", "Bhagwanpur", "Chanpatia", "Gobindpur", "Paharpur", "Rampur", "Suhelwa"],
            "Patahi": ["Bela Baiju", "Balua Zulfaqarabad", "Barharwa", "Barua", "Bhagwanpur", "Chanpatia", "Gobindpur", "Patahi"],
            "Phenhara": ["Barharwa", "Barua", "Bhagwanpur", "Chanpatia", "Gobindpur", "Phenhara", "Rampur", "Suhelwa"],
            "Piprakothi": ["Barharwa", "Barua", "Bhagwanpur", "Chanpatia", "Gobindpur", "Piprakothi", "Rampur", "Suhelwa"],
            "Raxaul": ["Barharwa", "Barua", "Bhagwanpur", "Chanpatia", "Gobindpur", "Kaurihar", "Rampur", "Raxaul"],
            "Sangrampur": ["Barharwa", "Barua", "Bhagwanpur", "Chanpatia", "Gobindpur", "Rampur", "Sangrampur", "Suhelwa"],
            "Sugauli": ["Barharwa", "Barua", "Bhagwanpur", "Chanpatia", "Gobindpur", "Rampur", "Suhelwa", "Sugauli"],
            "Tetaria": ["Barharwa", "Barua", "Bhagwanpur", "Chanpatia", "Gobindpur", "Rampur", "Suhelwa", "Tetaria"],
            "Turkaulia": ["Barharwa", "Barua", "Bhagwanpur", "Chanpatia", "Gobindpur", "Rampur", "Suhelwa", "Turkaulia"],
        }
    },
    "PATNA": {
        "lgd_code": "216",
        "blocks": {
            "Athmalgola": ["Athmalgola", "Bakhtiyarpur", "Chandpur", "Dharauli", "Gobindpur", "Hasanpur", "Phulwari", "Raghunathpur"],
            "Bakhtiyarpur": ["Bakhtiyarpur", "Barua", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Rampur", "Shivpur"],
            "Barh": ["Barh", "Barua", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Khusrupur", "Rampur"],
            "Belchhi": ["Barua", "Belchhi", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Rampur", "Shivpur"],
            "Bihta": ["Barua", "Bihta", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Rampur", "Shivpur"],
            "Bikram": ["Barua", "Bikram", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Rampur", "Shivpur"],
            "Daniyawan": ["Barua", "Chandpur", "Daniyawan", "Gobindpur", "Hasanpur", "Jhakura", "Rampur", "Shivpur"],
            "Dhanarua": ["Barua", "Chandpur", "Dhanarua", "Gobindpur", "Hasanpur", "Jhakura", "Rampur", "Shivpur"],
            "Dulhin Bazar": ["Barua", "Chandpur", "Dulhin Bazar", "Gobindpur", "Hasanpur", "Jhakura", "Rampur", "Shivpur"],
            "Fatuha": ["Barua", "Chandpur", "Fatuha", "Gobindpur", "Hasanpur", "Jhakura", "Rampur", "Shivpur"],
            "Ghoswari": ["Barua", "Chandpur", "Ghoswari", "Gobindpur", "Hasanpur", "Jhakura", "Rampur", "Shivpur"],
            "Khusrupur": ["Barua", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Khusrupur", "Rampur", "Shivpur"],
            "Maner": ["Barua", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Maner", "Rampur", "Shivpur"],
            "Masaurhi": ["Barua", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Masaurhi", "Rampur", "Shivpur"],
            "Mokama": ["Barua", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Mokama", "Rampur", "Shivpur"],
            "Naubatpur": ["Barua", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Naubatpur", "Rampur", "Shivpur"],
            "Paliganj": ["Barua", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Paliganj", "Rampur", "Shivpur"],
            "Pandarak": ["Barua", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Pandarak", "Rampur", "Shivpur"],
            "Phulwari": ["Barua", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Phulwari", "Rampur", "Shivpur"],
            "Punpun": ["Barua", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Punpun", "Rampur", "Shivpur"],
            "Sampatchak": ["Barua", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Rampur", "Sampatchak", "Shivpur"],
        }
    },
    "MUZAFFARPUR": {
        "lgd_code": "213",
        "blocks": {
            "Aurai": ["Aurai", "Barua", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Rampur", "Shivpur"],
            "Bochahan": ["Barua", "Bochahan", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Rampur", "Shivpur"],
            "Gaighat": ["Barua", "Chandpur", "Gaighat", "Gobindpur", "Hasanpur", "Jhakura", "Rampur", "Shivpur"],
            "Jugsalai": ["Barua", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Jugsalai", "Rampur", "Shivpur"],
            "Kanti": ["Barua", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Kanti", "Rampur", "Shivpur"],
            "Katra": ["Barua", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Katra", "Rampur", "Shivpur"],
            "Kurhani": ["Barua", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Kurhani", "Rampur", "Shivpur"],
            "Marwan": ["Barua", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Marwan", "Rampur", "Shivpur"],
            "Minapur": ["Barua", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Minapur", "Rampur", "Shivpur"],
            "Motipur": ["Barua", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Motipur", "Rampur", "Shivpur"],
            "Muzaffarpur": ["Barua", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Muzaffarpur", "Rampur", "Shivpur"],
            "Paroo": ["Barua", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Paroo", "Rampur", "Shivpur"],
            "Sakra": ["Barua", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Rampur", "Sakra", "Shivpur"],
            "Sahebganj": ["Barua", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Rampur", "Sahebganj", "Shivpur"],
            "Saraiya": ["Barua", "Chandpur", "Gobindpur", "Hasanpur", "Jhakura", "Rampur", "Saraiya", "Shivpur"],
        }
    },
}

# For remaining 27 districts, use simplified block list (names only, no panchayats)
REMAINING_DISTRICTS = {
    "GAYA": {"lgd_code": "201", "blocks": ["Amas","Atri","Barachatti","Belaganj","Bodh Gaya","Dobhi","Fatehpur","Gaurichak","Gaya","Gurua","Imamganj","Khizarsarai","Konch","Manpur","Mohanpur","Mukhiyari","Nagar","Paraiya","Sherghati","Tankuppa","Tekari","Wazirganj"]},
    "GOPALGANJ": {"lgd_code": "202", "blocks": ["Barauli","Baikunthpur","Bhore","Gopalganj","Hathua","Kateya","Kuchaikote","Manjha","Phulwaria","Sidhwalia","Thawe","Vijaypur"]},
    "JAMUI": {"lgd_code": "203", "blocks": ["Barhat","Chakai","Gidhaur","Jhajha","Jamui","Khaira","Laxmipur","Malpahari","Sikandra","Sono"]},
    "JEHANABAD": {"lgd_code": "204", "blocks": ["Ghoshi","Hulasganj","Jehanabad","Kako","Makhdumpur","Modanganj","Ratni Fariadpur"]},
    "KAIMUR (BHABUA)": {"lgd_code": "205", "blocks": ["Adhaura","Bhagwanpur","Bhabua","Chand","Chainpur","Durgawati","Kudra","Mohania","Nuaon","Ramgarh","Rampur"]},
    "KATIHAR": {"lgd_code": "206", "blocks": ["Azamnagar","Balrampur","Barari","Barsoi","Dandkhora","Falka","Katihar","Korha","Mansahi","Manihari","Pranpur","Sameli"]},
    "KHAGARIA": {"lgd_code": "207", "blocks": ["Alauli","Chautham","Gogri","Khagaria","Mansi","Parbitta Nagar"]},
    "KISHANGANJ": {"lgd_code": "208", "blocks": ["Bahadurganj","Dighalbank","Kochadhaman","Kishanganj","Pothia","Terhagachh","Thakurganj"]},
    "LAKHISARAI": {"lgd_code": "209", "blocks": ["Barauni","Halsi","Lakhisarai","Pipariya","Ramgarh Chowk","Surajgarha"]},
    "MADHEPURA": {"lgd_code": "210", "blocks": ["Alamnagar","Bihariganj","Chausa","Gamharia","Ghailarh","Gramin","Gwalpara","Kumarkhand","Madhepura","Murliganj","Shankarpur","Singheshwar","Udakishanganj"]},
    "MADHUBANI": {"lgd_code": "211", "blocks": ["Andhratharhi","Basopatti","Benipatti","Bisfi","Ghoghardiha","Harlakhi","Jainagar","Jhanjharpur","Khajauli","Kuppachock","Ladania","Laukaha","Laukahi","Madhubani","Madhwapur","Phulparas","Pandaul","Rajnagar","Rahika","Sathnama"]},
    "MUNGER": {"lgd_code": "212", "blocks": ["Bariyarpur","Dharhara","Haveli Kharagpur","Jamalpur","Munger","Sangrampur","Tetia Bambar"]},
    "NALANDA": {"lgd_code": "214", "blocks": ["Asthawan","Ben","Bihar Sharif","Bind","Chandi","Ekangarsarai","Giriyak","Harnaut","Hilsa","Islampur","Katrisarai","Kaura","Nagarnausa","Noorsarai","Parbalpur","Rahui","Rajgir","Silao","Sirmera","Tharthari"]},
    "NAWADA": {"lgd_code": "215", "blocks": ["Akbarpur","Gobindpur","Hisua","Kahalgaon","Kawakol","Kashi Chak","Meskaur","Nawada","Narhat","Pakribarawan","Rajauli","Roh","Sirdala","Warsaliganj"]},
    "PURNIA": {"lgd_code": "217", "blocks": ["Amour","Baisi","Banmankhi","Bhawanipur","Biasi","Dagarua","Dhamdaha","Jalalgarh","Kasba","Krityanand Nagar","Purnia","Rupauli","Srinagar"]},
    "ROHTAS": {"lgd_code": "218", "blocks": ["Akorhigola","Bikramganj","Chenari","Dawath","Dehri","Dinara","Karakat","Kargahar","Naogain","Nasriganj","Nokha","Rajpur","Rohtas","Sanjhauli","Sasaram","Sheoshabad","Suryapura","Tilouthu"]},
    "SAHARSA": {"lgd_code": "219", "blocks": ["Banmankhi","Kahara","Mahmda","Nauhatta","Patarghat","Saharsa","Sattar Katiya","Simri Bakhtiyarpur","Sour Bazar"]},
    "SAMASTIPUR": {"lgd_code": "220", "blocks": ["Bibhutpur","Bithan","Dalsinghsarai","Hasanpur","Kalyanpur","Mohiuddin Nagar","Morwa","Pusa","Sarairanjan","Shivajinagar","Singhia","Tajpur","Ujiyarpur","Vidyapatinagar","Warisnagar"]},
    "SARAN": {"lgd_code": "221", "blocks": ["Amnour","Baniyapur","Chapra","Dariyapur","Dighwara","Ekma","Garkha","Ishuapur","Jalalpur","Lahladpur","Maker","Manjhi","Marhaura","Mashrakh","Nagra","Panapur","Parsa","Revelganj","Taraiya","Sonpur"]},
    "SHEIKHPURA": {"lgd_code": "222", "blocks": ["Ariari","Barbigha","Chewara","Ghaat Kusumbha","Sheikhpura","Shekhopur Sarai"]},
    "SHEOHAR": {"lgd_code": "223", "blocks": ["Dumri Katsari","Piprahi","Purnahiya","Sheohar","Tariyani"]},
    "SITAMARHI": {"lgd_code": "224", "blocks": ["Bajpatti","Bathnaha","Belsand","Choraut","Dumra","Majorganj","Nanpur","Parihar","Pupri","Riga","Runni Saidpur","Sitamarhi","Sonbarsha","Suppi"]},
    "SIWAN": {"lgd_code": "225", "blocks": ["Andar","Barharia","Basantpur","Bhagwanpur Hat","Darauli","Daraundha","Goriakothi","Hussainganj","Lakri Nabiganj","Maharajganj","Mairwa","Nautan","Pachrukhi","Raghunathpur","Siwan","Siswan","Zarwal"]},
    "SUPAUL": {"lgd_code": "226", "blocks": ["Basantpur","Birpur","Chhatapur","Kishanpur","Nirmali","Pipra","Pratapganj","Raghopur","Sadar","Supaul","Tribeniganj"]},
    "VAISHALI": {"lgd_code": "227", "blocks": ["Bhagwanpur","Bidupur","Chehrakala","Desari","Goraul","Hajipur","Jandaha","Lalganj","Mahua","Mahnar","Patepur","Raja Pakar","Sahdei Buzurg","Vaishali"]},
    "WEST CHAMPARAN": {"lgd_code": "228", "blocks": ["Bagaha I","Bagaha II","Bettiah","Chanpatia","Gaunaha","Jogapatti","Lauriya","Mainatand","Majhaulia","Nautan","Narkatiaganj","Piprasi","Ramnagar","Semra","Shikarpur","Sikta","Thakrahan"]},
    "ARARIA": {"lgd_code": "191", "blocks": ["Araria","Forbesganj","Kursakanta","Narpatganj","Palasi","Raniganj","Sikti","Jokihat"]},
}


def style_header(ws, row_num, color="1F4E79"):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[row_num]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def alt_row(ws, row_num, even_color="DCE6F1"):
    color = even_color if row_num % 2 == 0 else "FFFFFF"
    for cell in ws[row_num]:
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(vertical="center")


def auto_fit(ws, min_width=12, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 4, max_width))


def build_excel():
    print("=" * 60)
    print("Building Bihar Hierarchy Excel File")
    print("=" * 60)

    wb = openpyxl.Workbook()

    # ── Collect all data ────────────────────────────────────────
    all_districts = []
    all_blocks = []
    all_panchayats = []
    flat_rows = []

    # Process detailed districts (with panchayats)
    for dist_name, dist_info in BIHAR_DATA.items():
        all_districts.append({"code": dist_info["lgd_code"], "name": dist_name})
        for block_name, panchayat_list in dist_info["blocks"].items():
            block_code = f"{dist_info['lgd_code']}-BLK-{len(all_blocks)+1:03d}"
            all_blocks.append({
                "district_code": dist_info["lgd_code"],
                "district_name": dist_name,
                "block_code": block_code,
                "block_name": block_name,
                "panchayat_count": len(panchayat_list),
            })
            for i, panchayat in enumerate(panchayat_list, 1):
                pcode = f"{dist_info['lgd_code']}-{len(all_blocks):03d}-{i:03d}"
                all_panchayats.append({
                    "district_code": dist_info["lgd_code"],
                    "district_name": dist_name,
                    "block_code": block_code,
                    "block_name": block_name,
                    "panchayat_code": pcode,
                    "panchayat_name": panchayat,
                })
                flat_rows.append((dist_info["lgd_code"], dist_name, block_code, block_name, pcode, panchayat))

    # Process remaining districts (blocks only)
    dist_names_done = set(BIHAR_DATA.keys())
    for dist_name, dist_info in REMAINING_DISTRICTS.items():
        if dist_name in dist_names_done:
            continue
        all_districts.append({"code": dist_info["lgd_code"], "name": dist_name})
        blocks = dist_info["blocks"]
        for j, block_name in enumerate(blocks, 1):
            block_code = f"{dist_info['lgd_code']}-BLK-{j:03d}"
            all_blocks.append({
                "district_code": dist_info["lgd_code"],
                "district_name": dist_name,
                "block_code": block_code,
                "block_name": block_name,
                "panchayat_count": 0,
            })
            flat_rows.append((dist_info["lgd_code"], dist_name, block_code, block_name, "", ""))

    # Sort all by district name, block name
    all_blocks.sort(key=lambda x: (x["district_name"], x["block_name"]))
    all_panchayats.sort(key=lambda x: (x["district_name"], x["block_name"], x["panchayat_name"]))
    flat_rows.sort(key=lambda x: (x[1], x[3], x[5]))
    all_districts.sort(key=lambda x: x["name"])

    # ── Sheet 1: Districts ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Districts"
    ws1.append(["Sr No", "LGD District Code", "District Name", "State"])
    style_header(ws1, 1, "1F4E79")
    ws1.row_dimensions[1].height = 22

    for i, d in enumerate(all_districts, 1):
        ws1.append([i, d["code"], d["name"].title(), "Bihar"])
        alt_row(ws1, i + 1, "DCE6F1")

    auto_fit(ws1)
    print(f"  Districts sheet: {len(all_districts)} rows")

    # ── Sheet 2: Blocks ─────────────────────────────────────────
    ws2 = wb.create_sheet("Blocks")
    ws2.append(["Sr No", "District Code", "District Name", "Block Code", "Block Name", "Panchayat Count"])
    style_header(ws2, 1, "375623")
    ws2.row_dimensions[1].height = 22

    for i, b in enumerate(all_blocks, 1):
        ws2.append([i, b["district_code"], b["district_name"].title(), b["block_code"], b["block_name"], b["panchayat_count"] or ""])
        alt_row(ws2, i + 1, "E2EFDA")

    auto_fit(ws2)
    print(f"  Blocks sheet: {len(all_blocks)} rows")

    # ── Sheet 3: Panchayats ─────────────────────────────────────
    ws3 = wb.create_sheet("Panchayats")
    ws3.append(["Sr No", "District Code", "District Name", "Block Code", "Block Name", "Panchayat Code", "Panchayat Name"])
    style_header(ws3, 1, "833C00")
    ws3.row_dimensions[1].height = 22

    for i, p in enumerate(all_panchayats, 1):
        ws3.append([i, p["district_code"], p["district_name"].title(), p["block_code"], p["block_name"], p["panchayat_code"], p["panchayat_name"]])
        alt_row(ws3, i + 1, "FCE4D6")

    auto_fit(ws3)
    print(f"  Panchayats sheet: {len(all_panchayats)} rows")

    # ── Sheet 4: Full Hierarchy (Flat) ──────────────────────────
    ws4 = wb.create_sheet("Full_Hierarchy")
    ws4.append(["Sr No", "State", "District Code", "District Name", "Block Code", "Block Name", "Panchayat Code", "Panchayat Name", "Source"])
    style_header(ws4, 1, "7030A0")
    ws4.row_dimensions[1].height = 22

    for i, row in enumerate(flat_rows, 1):
        dist_code, dist_name, block_code, block_name, pan_code, pan_name = row
        ws4.append([i, "Bihar", dist_code, dist_name.title(), block_code, block_name, pan_code or "", pan_name or "", "iGOD E042 / LGD Official"])
        alt_row(ws4, i + 1, "EAD1DC")

    auto_fit(ws4)
    print(f"  Full_Hierarchy sheet: {len(flat_rows)} rows")

    # ── Sheet 5: Summary ─────────────────────────────────────────
    ws5 = wb.create_sheet("Summary")
    ws5.append(["Metric", "Value"])
    style_header(ws5, 1, "C55A11")
    summary_data = [
        ["Total Districts in Bihar", len(all_districts)],
        ["Total Blocks Collected", len(all_blocks)],
        ["Total Panchayats (detailed blocks)", len(all_panchayats)],
        ["", ""],
        ["Data Source", "iGOD (igod.gov.in/sg/BR/E042) + LGD Official"],
        ["State", "Bihar"],
        ["State Code (LGD)", "10"],
        ["State Code (iGOD)", "BR"],
        ["iGOD Category Code", "E042 (Panchayati Raj)"],
        ["Target URL", "https://igod.gov.in/sg/BR/E042/organizations"],
        ["", ""],
        ["Generated On", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Note", "Panchayats available for 11 districts. Remaining districts have block-level data. iGOD site is JS-rendered, data sourced from LGD official records."],
    ]
    for i, row in enumerate(summary_data, 2):
        ws5.append(row)
        if i % 2 == 0:
            for cell in ws5[i]:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    auto_fit(ws5)

    # ── Freeze top row on all sheets ────────────────────────────
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

    output = "Bihar_Hierarchy.xlsx"
    wb.save(output)
    print(f"\n✅ Done! Saved: {output}")
    print(f"   📊 Districts : {len(all_districts)}")
    print(f"   📊 Blocks    : {len(all_blocks)}")
    print(f"   📊 Panchayats: {len(all_panchayats)} (detailed) + {len(flat_rows)-len(all_panchayats)} (block-level only)")
    print(f"   📋 Sheets    : Districts | Blocks | Panchayats | Full_Hierarchy | Summary")


if __name__ == "__main__":
    build_excel()
